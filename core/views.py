from django.shortcuts import render_to_response, HttpResponseRedirect
from django.core.context_processors import csrf

from workpacks.models import Workpack
from materials.models import Lineclass11011

from openpyxl import load_workbook


def homepageview(request):
    context = dict()
    context.update(csrf(request))

    context['workpacks'] = Workpack.objects.all()

    context['lineclass11011'] = Lineclass11011.objects.all().distinct()

    return render_to_response('index.html', context)


def exportbom(request):
    bom_template = load_workbook(filename='exceltemplate.xlsx')
    ws = bom_template.active
    ws['C20'] = 'Hello'
    bom_template.save(filename='exceltemplate.xlsx')
    return HttpResponseRedirect('/home/')